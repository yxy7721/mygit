# -*- coding: utf-8 -*-
"""
Created on Tue Oct 25 14:26:16 2022

@author: yangxy
"""


import numpy as np
import pandas as pd
import xlwings as xw
import openpyxl as op


#读取各个基金的重仓股是什么
wb=op.load_workbook(r'D:\desktop\winddataget\【320w待取数】adjfac_hk.xlsx',data_only=True)
#wb=op.load_workbook(r'D:\desktop\winddataget\重仓股取数be.xlsx',data_only=False) 想要包含公式就搞这个
#wb[wb.sheetnames[0]].title
wbnew=op.Workbook()

for she in range(len(wb.sheetnames)):
    if wb[wb.sheetnames[she]].title=="Sheet1":
        break
    print(wb[wb.sheetnames[she]].title)
    if she==0:
        shenew=wbnew.active
        shenew.title=wb[wb.sheetnames[she]].title
    else:
        shenew=wbnew.create_sheet(wb[wb.sheetnames[she]].title)
    #df=pd.DataFrame(None,index=range(wb[wb.sheetnames[she]].max_row),columns=range(wb[wb.sheetnames[she]].max_column))
    for i in range(wb[wb.sheetnames[she]].max_row):
        for j in range(wb[wb.sheetnames[she]].max_column):
            shenew.cell(i+1,j+1).value=wb[wb.sheetnames[she]].cell(i+1,j+1).value
wb.close()
wbnew.save(r'D:\desktop\winddataget\hkadj.xlsx')
wbnew.close()


'''
sheet=wb[wb.sheetnames[0]]
sheet.cell(1,2).value
#wb[wb.sheetnames[0]]['B2'].value
sheet.max_row
sheet.max_column
for i in wb[wb.sheetnames[0]]:
    print(i)


#读取各个基金的重仓股的占比
wb=op.load_workbook(r'D:\desktop\indus_rolling\【已经】重仓占比取数.xlsx',data_only=True)
#wb[wb.sheetnames[0]].title
zcrateshtlis=[]
for she in wb:
    if she.title=="Sheet1":
        break
    print(she.title)
    df=pd.DataFrame(None,index=range(she.max_row),columns=range(she.max_column))
    for i in range(she.max_row):
        for j in range(she.max_column):
            df.iat[i,j]=she.cell(i+1,j+1).value
    zcrateshtlis.append(df)
wb.close()

funds=pd.DataFrame(zcrateshtlis[0].iloc[0,1:].copy())
funds.index=range(100)

#读取各个基金的所属行业
wb=op.load_workbook(r'D:\desktop\indus_rolling\code2indus.xlsx',data_only=True)
she=wb[wb.sheetnames[0]]
code2indus=pd.DataFrame(None,index=range(she.max_row),columns=range(she.max_column))
for i in range(she.max_row):
    for j in range(she.max_column):
        code2indus.iat[i,j]=she.cell(i+1,j+1).value
wb.close()

#读取各个基金所属行业（港股）
wb=op.load_workbook(r'D:\desktop\indus_rolling\code2indushk.xlsx',data_only=True)
she=wb[wb.sheetnames[0]]
code2indushk=pd.DataFrame(None,index=range(she.max_row),columns=range(she.max_column))
for i in range(she.max_row):
    for j in range(she.max_column):
        code2indushk.iat[i,j]=she.cell(i+1,j+1).value
wb.close()

code2indus=code2indus.drop([0],axis=0)
code2indushk=code2indushk.drop([0],axis=0)
code2indus=pd.concat([code2indus,code2indushk],axis=0)
code2indus.columns=['code','industry']
code2indus.index=range(len(code2indus.index))

wb=op.load_workbook(r'D:\desktop\indus_rolling\indusname.xlsx',data_only=True)
she=wb[wb.sheetnames[0]]
indusname=pd.DataFrame(None,index=range(she.max_row),columns=range(she.max_column))
for i in range(she.max_row):
    for j in range(she.max_column):
        indusname.iat[i,j]=she.cell(i+1,j+1).value
wb.close()
indusname=indusname.drop([0],axis=0)
indusname.columns=['induscode','name']

wb=op.load_workbook(r'D:\desktop\indus_rolling\indusret.xlsx',data_only=True)
she=wb[wb.sheetnames[0]]
indusret=pd.DataFrame(None,index=range(she.max_row),columns=range(she.max_column))
for i in range(she.max_row):
    for j in range(she.max_column):
        indusret.iat[i,j]=she.cell(i+1,j+1).value
wb.close()

#indusret.iloc[2:,1:].rank(axis=1,method="average",ascending=False)

#计算各个基金每个行业占基金股票投资的比重
code2indus0=code2indus.applymap(lambda x: x[:len(x)-3] if not(x is None) else None)
indusret.columns=indusret.iloc[0,:]
indusret.drop([indusret.index[0]],axis=0,inplace=True)
indusret.index=indusret.iloc[:,0]
indusret.drop([indusret.index[0]],axis=0,inplace=True)
indusret.drop([indusret.columns[0]],axis=1,inplace=True)
indusretrank=indusret.rank(axis=1,method="average",ascending=False)

for diji in range(10):
    sh=zcshtlis[diji].copy()
    sh.columns=sh.iloc[0,:]
    sh.drop([sh.index[0]],axis=0,inplace=True)
    sh.index=sh.iloc[:,0]
    sh.drop([sh.index[0]],axis=0,inplace=True)
    sh.drop([sh.columns[0]],axis=1,inplace=True)
    zcshtlis[diji]=sh.copy()
    
    sh=zcrateshtlis[diji].copy()
    sh.columns=sh.iloc[0,:]
    sh.drop([sh.index[0]],axis=0,inplace=True)
    sh.index=sh.iloc[:,0]
    sh.drop([sh.index[0]],axis=0,inplace=True)
    sh.drop([sh.columns[0]],axis=1,inplace=True)
    zcrateshtlis[diji]=sh.copy()
    
defen=pd.DataFrame(index=zcshtlis[0].index,columns=zcshtlis[0].columns)
defen.iloc[0,:]=[0 for i in range(len(defen.columns))]
induschange=defen.copy()
daibiao=defen.copy()
for i in range(len(zcshtlis[0].columns)):
    lagspread=pd.DataFrame(0.00,index=indusret.columns,columns=['rate'])
    for dateindex in range(0,len(indusret.index)):
        
        induspack=pd.DataFrame(0.00,index=indusret.columns,columns=['rate'])
        indusspread=pd.DataFrame(0.00,index=indusret.columns,columns=['rate'])
        for diji in range(len(zcshtlis)):
            piao=zcshtlis[diji].loc[zcshtlis[diji].index[dateindex],zcshtlis[diji].columns[i]]
            if len(code2indus0[code2indus0['code']==piao].index)==0:
                continue
            hangye=code2indus0[code2indus0['code']==piao].iat[0,1]
            hangye=hangye+".SI"
            
            indusspread.loc[hangye,:][0]=indusspread.loc[hangye,][0]+zcrateshtlis[diji].iat[dateindex,i]
            induspack.loc[hangye,:][0]=induspack.loc[hangye,][0]+indusretrank.loc[indusret.index[dateindex],hangye]*zcrateshtlis[diji].iat[dateindex,i]
        if dateindex==0:
            induschange.iat[dateindex,i]=None
        else:
            induschange.iat[dateindex,i]=((lagspread-indusspread)**2).sum()[0]
        lagspread=indusspread.copy()
        defen.iat[dateindex,i]=induspack.sum()[0]
    
defen.to_excel(r'D:\desktop\indus_rolling\defennew.xlsx',header=True,index=True)
induschange.to_excel(r'D:\desktop\indus_rolling\induschangenew.xlsx',header=True,index=True)

indusrate_ex=pd.DataFrame(0.00,index=indusret.columns,columns=['rate'])
for i in range(len(zcshtlis[0].columns)):
    for dateindex in range(1,len(indusret.index)):
        
        indusrate_now=pd.DataFrame(0.00,index=indusret.columns,columns=['rate']).copy()
        for diji in range(len(zcshtlis)):
            piao=zcshtlis[diji].loc[zcshtlis[diji].index[dateindex],zcshtlis[diji].columns[i]]
            if len(code2indus0[code2indus0['code']==piao].index)==0:
                continue
            hangye=code2indus0[code2indus0['code']==piao].iat[0,1]
            hangye=hangye+".SI"
            indusrate_now.loc[hangye,:][0]=indusrate_now.loc[hangye,:][0]+zcrateshtlis[diji].iat[dateindex,i]
        zuidanew=indusrate_now[indusrate_now['rate']==indusrate_now.max()[0]].index[0] if indusrate_now.max()[0]>0 else None
        zuidaold=indusrate_ex[indusrate_ex['rate']==indusrate_ex.max()[0]].index[0] if indusrate_ex.max()[0]>0 else None
        if (zuidanew is None) or (zuidaold is None):
            pass
        elif zuidanew!=zuidaold:
            huancangstr=indusrate_ex[indusrate_ex['rate']==indusrate_ex.max()[0]].index[0]
            huancangstr=indusname[indusname['induscode']==huancangstr].iat[0,1]
            miaoshu="由重仓"+huancangstr+"行业调仓为"
            huancangstr=indusrate_now[indusrate_now['rate']==indusrate_now.max()[0]].index[0]
            huancangstr=indusname[indusname['induscode']==huancangstr].iat[0,1]
            miaoshu=miaoshu+huancangstr+"行业"
            daibiao.iat[dateindex,i]=miaoshu
        indusrate_ex=indusrate_now.copy()

daibiao.to_excel(r'D:\desktop\indus_rolling\daibiaonew.xlsx',header=True,index=True)

#程序外整理defennew后，后续调试可从此开始
wb=op.load_workbook(r'D:\desktop\indus_rolling\defen.xlsx',data_only=True)
she=wb[wb.sheetnames[0]]
defen=pd.DataFrame(None,index=range(she.max_row),columns=range(she.max_column))
for i in range(she.max_row):
    for j in range(she.max_column):
        defen.iat[i,j]=she.cell(i+1,j+1).value
wb.close()
sh=defen.copy()
sh.columns=sh.iloc[0,:]
sh.drop([sh.index[0]],axis=0,inplace=True)
sh.index=sh.iloc[:,0]
sh.drop([sh.columns[0]],axis=1,inplace=True)
defen=sh.copy()

wb=op.load_workbook(r'D:\desktop\indus_rolling\daibiao.xlsx',data_only=True)
she=wb[wb.sheetnames[0]]
daibiao=pd.DataFrame(None,index=range(she.max_row),columns=range(she.max_column))
for i in range(she.max_row):
    for j in range(she.max_column):
        daibiao.iat[i,j]=she.cell(i+1,j+1).value
wb.close()
sh=daibiao.copy()
sh.columns=sh.iloc[0,:]
sh.drop([sh.index[0]],axis=0,inplace=True)
sh.index=sh.iloc[:,0]
sh.drop([sh.columns[0]],axis=1,inplace=True)
daibiao=sh.copy()

wb=op.load_workbook(r'D:\desktop\indus_rolling\induschange.xlsx',data_only=True)
she=wb[wb.sheetnames[0]]
induschange=pd.DataFrame(None,index=range(she.max_row),columns=range(she.max_column))
for i in range(she.max_row):
    for j in range(she.max_column):
        induschange.iat[i,j]=she.cell(i+1,j+1).value
sh=induschange.copy()
sh.columns=sh.iloc[0,:]
sh.drop([sh.index[0]],axis=0,inplace=True)
sh.index=sh.iloc[:,0]
sh.drop([sh.columns[0]],axis=1,inplace=True)
induschange=sh.copy()
induschange=induschange.applymap(lambda x:None if x==0 else x)

huizong=pd.DataFrame(None,columns=defen.columns,index=['Rx','Rchange','N','总得分','经典战役'])
defen=defen.applymap(lambda x:x if x!=0 else None).copy()

huizong.iloc[0,:]=defen.mean(axis=0,skipna=True).rank(axis=0,method="average",ascending=True).copy()
huizong.iloc[1,:]=induschange.mean(axis=0,skipna=True).rank(axis=0,method="average",ascending=False).copy()
huizong.iloc[0:2,:]=huizong.iloc[0:2,:]/len(huizong.columns)
for i in range(len(huizong.columns)):
    if defen.count(axis=0)[i]<=4:
        huizong.iat[2,i]=0.4
    elif defen.count(axis=0)[i]<=8:
        huizong.iat[2,i]=0.3
    elif defen.count(axis=0)[i]<=12:
        huizong.iat[2,i]=0.2
    elif defen.count(axis=0)[i]<=16:
        huizong.iat[2,i]=0.1
    else:
        huizong.iat[2,i]=0

huizong.iloc[3,:]=huizong.iloc[0,:]*0.7+huizong.iloc[1,:]*0.3+huizong.iloc[2,:]

for i in range(len(daibiao.columns)):
    tiaocangquan=""
    for j in range(len(daibiao.index)):
        if (daibiao.iat[j,i] is None) or (daibiao.iat[j,i]==0):
            pass
        else:
            tiaocangquan=tiaocangquan+daibiao.index[j].strftime('%Y-%m')+daibiao.iat[j,i]+"\n"
    if len(tiaocangquan)>0:
        tiaocangquan=tiaocangquan[:len(tiaocangquan)-3]
    huizong.loc[huizong.index[4],daibiao.columns[i]]=tiaocangquan
        

huizong.T.to_excel(r'D:\desktop\indus_rolling\huizongnew.xlsx',header=True,index=True)
'''