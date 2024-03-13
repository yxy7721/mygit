# -*- coding: utf-8 -*-
"""
Created on Wed Jul 20 09:58:39 2022

@author: yangxy
"""

import numpy as np
import pandas as pd
import xlwings as xw
import datetime as dt
import openpyxl as op

app=xw.App(visible=False,add_book=False)
app.display_alerts=False
app.screen_updating=False

'''
hang=len(wb.sheets[0].range('A1').current_region.rows)
hang
lie=len(wb.sheets[0].range('A1').current_region.columns)
#df=wb.sheets[0].range((1,1),(hang,lie)).options(pd.DataFrame,index=False).value
wb.sheets[0].range((7,1)).formula='=A2+B2'
wb.sheets[0].range((7,1)).value
wb.sheets[0].range((7,1)).formula
'''

wb=app.books.open(r'D:\desktop\winddataget\pool.xlsx')
fundpool=wb.sheets[0].range((1,1)).options(pd.DataFrame,index=False,expand='table').value
timepool=wb.sheets[1].range((1,1)).options(pd.DataFrame,index=False,expand='table').value
wb.close()

kk=[timepool.iat[i,0].strftime("%Y-%m-%d") for i in range(len(timepool.index))]
#kk=[kk[0],kk[1]].copy()
#fundpool=fundpool.iloc[:3,:].copy()
dataget=pd.DataFrame(columns=fundpool.iloc[:,0],index=kk)
#dataget.resample('3M',axis=0,closed="right",label="right").last()

#一、取重仓股代码
shuchu=app.books.add()
for sh in range(10):
    print(sh)
    dataget=pd.DataFrame(columns=fundpool.iloc[:,0],index=kk)
    for i in range(len(dataget.index)):
        for j in range(len(dataget.columns)):
            dataget.iat[i,j]=str(i+2)+","+str(j+2)
    dataget=dataget.applymap(lambda x: '=f_prt_topstockcode(indirect(address(1,{},4)),indirect(address({},1,4)),{})'.format(x.split(',')[1],x.split(',')[0],sh+1))
    dataget.index.name='date'
    shuchu.sheets.add()
    shuchu.sheets[0].name="第"+str(sh+1)+"大重仓股代码"
    shuchu.sheets[0].range('A1').options(pd.DataFrame, index=True).value=dataget

shuchu.save(r'D:\desktop\winddataget\重仓股取数.xlsx')

#二、取重仓占比
shuchu=app.books.add()
for sh in range(10):
    print(sh)
    dataget=pd.DataFrame(columns=fundpool.iloc[:,0],index=kk)
    for i in range(len(dataget.index)):
        for j in range(len(dataget.columns)):
            dataget.iat[i,j]=str(i+2)+","+str(j+2)
    dataget=dataget.applymap(lambda x: '=f_prt_heavilyheldstocktonav(indirect(address(1,{},4)),indirect(address({},1,4)),{})'.format(x.split(',')[1],x.split(',')[0],sh+1))
    dataget.index.name='date'
    shuchu.sheets.add()
    shuchu.sheets[0].name="第"+str(sh+1)+"大重仓股占比"
    shuchu.sheets[0].range('A1').options(pd.DataFrame, index=True).value=dataget
shuchu.save(r'D:\desktop\winddataget\重仓占比取数.xlsx')

#三、取收盘价
wbnew=op.Workbook()
shenew=wbnew.active

for i in range(len(kk)+1):
    for j in range(len(fundpool.iloc[:,0])+1):
        if i==0 and j==0:
            shenew.cell(i+1,j+1).value="date"
        elif i==0 and j!=0:
            shenew.cell(i+1,j+1).value=fundpool.iloc[:,0][j-1]
        elif i!=0 and j==0:
            shenew.cell(i+1,j+1).value=kk[i-1]
        else:
            tmp1=fundpool.iloc[:,0][j-1]
            tmp2=kk[i-1]
            print(tmp1,"___",tmp2)
            #dataget.iat[i,j]='=f_nav_accumulated(indirect(address(1,{},4)),indirect(address({},1,4)))'.format(x.split(',')[1],x.split(',')[0])
            shenew.cell(i+1,j+1).value=(
                "=s_dq_close(\"{}\",\"{}\",1)".format(tmp1,tmp2)
                )
wbnew.save(r'D:\desktop\winddataget\hkclose.xlsx')
wbnew.close()

#四、取复权因子
wbnew=op.Workbook()
shenew=wbnew.active

for i in range(len(kk)+1):
    for j in range(len(fundpool.iloc[:,0])+1):
        if i==0 and j==0:
            shenew.cell(i+1,j+1).value="date"
        elif i==0 and j!=0:
            shenew.cell(i+1,j+1).value=fundpool.iloc[:,0][j-1]
        elif i!=0 and j==0:
            shenew.cell(i+1,j+1).value=kk[i-1]
        else:
            tmp1=fundpool.iloc[:,0][j-1]
            tmp2=kk[i-1]
            print(tmp1,"___",tmp2)
            #dataget.iat[i,j]='=f_nav_accumulated(indirect(address(1,{},4)),indirect(address({},1,4)))'.format(x.split(',')[1],x.split(',')[0])
            shenew.cell(i+1,j+1).value=(
                "=s_dq_adjfactor2(\"{}\",\"{}\")".format(tmp1,tmp2)
                )
wbnew.save(r'D:\desktop\winddataget\adjfac_hk.xlsx')
wbnew.close()

'''
shuchu=app.books.add()
dataget=pd.DataFrame(columns=fundpool.iloc[:,0],index=kk)
for i in range(len(dataget.index)):
    for j in range(len(dataget.columns)):
        break
        x=str(i+2)+","+str(j+2)
        print(x)
        dataget.iat[i,j]='=f_nav_accumulated(indirect(address(1,{},4)),indirect(address({},1,4)))'.format(x.split(',')[1],x.split(',')[0])
#dataget=dataget.applymap(lambda x: '=f_nav_accumulated(indirect(address(1,{},4)),indirect(address({},1,4)))'.format(x.split(',')[1],x.split(',')[0]))
dataget.index.name='date'
shuchu.sheets.add()
shuchu.sheets[0].name="基金净值"
shuchu.sheets[0].range('A1').options(pd.DataFrame, index=True).value=dataget
shuchu.save(r'D:\desktop\winddataget\基金净值取数0118.xlsx')

#f_nav_accumulated(Wind代码,交易日期)
'''

#app.quit()
app.kill()




























